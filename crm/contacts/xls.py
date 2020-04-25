import os
import urllib.parse
from openpyxl import Workbook, load_workbook
from .models import Contact
from django.db.models.functions import Lower
from django.http import HttpResponse


def import_contacts(file_name, user):
    file_name = urllib.parse.unquote(file_name)
    workbook = load_workbook(filename=file_name)
    sheet = workbook.active

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        contact = Contact(first_name=row[0], last_name=row[1])
        if row[2]:
            contact.email = row[2]
        if row[3]:
            contact.address_line_1 = row[3]
        if row[4]:
            contact.address_line_2 = row[4]
        if row[5]:
            contact.city = row[5]
        if row[6]:
            contact.state = row[6]
        if row[7]:
            contact.zipcode = row[7]
        contact.user = user
        contact.save()

    if os.path.isfile(file_name):
        os.remove(file_name)


def export_contacts(user):
    """
    Downloads all contacts as an Excel file with a single worksheet
    """
    contact_queryset = Contact.objects.filter(user=user).order_by(
        Lower('first_name'), Lower('last_name'))

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=contacts.xlsx'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Contacts'

    columns = [
        'First Name',
        'Last Name',
        'Email',
        'Address Line 1',
        'Address Line 2',
        'City',
        'State',
        'Zipcode',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for contact in contact_queryset:
        row_num += 1

        row = [
            contact.first_name,
            contact.last_name,
            contact.email,
            contact.address_line_1,
            contact.address_line_2,
            contact.city,
            contact.state,
            contact.zipcode
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response
