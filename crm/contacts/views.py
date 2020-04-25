from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from django.core.files.storage import FileSystemStorage
from django.conf import settings
from django.db.models.functions import Lower

import os
import urllib.parse

from openpyxl import Workbook, load_workbook

from .models import Contact
from .forms import ContactForm
from .postcards import send_postcard

@login_required(login_url="/login/")
def import_contacts(request, file_name):
    file_name = urllib.parse.unquote(file_name)
    workbook = load_workbook(filename=file_name)
    sheet = workbook.active

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        contact = Contact(first_name=row[0], last_name=row[1])
        if row[2]: contact.email = row[2]
        if row[3]: contact.address_line_1 = row[3]
        if row[4]: contact.address_line_2 = row[4]
        if row[5]: contact.city = row[5]
        if row[6]: contact.state = row[6]
        if row[7]: contact.zipcode = row[7]
        contact.user = request.user
        contact.save()

    if os.path.isfile(file_name):
       os.remove(file_name)

@login_required(login_url="/login/")
def export_contacts(request):
    """
    Downloads all contacts as an Excel file with a single worksheet
    """
    contact_queryset = Contact.objects.filter(user=request.user).order_by(Lower('first_name'), Lower('last_name'))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=contacts.xlsx'
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Contacts'

    columns = [
        'First Name',
        'Last Name',
        'Email',
        'Address Line 1',
        'Address Line 2',
        'City',
        'State',
        'Zipcode',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for contact in contact_queryset:
        row_num += 1

        row = [
            contact.first_name,
            contact.last_name,
            contact.email,
            contact.address_line_1,
            contact.address_line_2,
            contact.city,
            contact.state,
            contact.zipcode
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


@login_required(login_url="/login/")
def index(request):
    if request.method == 'POST' and 'document' in request.FILES:
        uploaded_file = request.FILES['document']
        fs = FileSystemStorage()
        name = fs.save(uploaded_file.name, uploaded_file)
        import_contacts(request, fs.url(name)[1:])
    contact_list = Contact.objects.filter(user=request.user).order_by(Lower('first_name'), Lower('last_name'))
    context = {'contact_list': contact_list}
    export = request.GET.get('export', False)
    if export:
        return export_contacts(request)
    return render(request, 'index.html', context)

@login_required(login_url="/login/")
def detail(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id)
    if request.method == 'POST':
        try:
            send_postcard(contact, request.POST['postcard_text'])
        except Exception as e:
            return render(request, 'pages/profile.html', {'contact': contact, 'postcard_error': str(e)})
        return render(request, 'pages/profile.html', {'contact': contact, 'postcard_success': True})
    return render(request, 'pages/profile.html', {'contact': contact})

@login_required(login_url="/login/")
def contact_create_view(request):
    form = ContactForm(request.POST or None, request.FILES)
    if form.is_valid():
        contact = form.save(commit=False)
        contact.user = request.user
        contact.save()
        return HttpResponseRedirect(reverse('contacts:detail', args=(contact.id,)))

    return render(request, 'pages/contact_create.html', {'form': form})

def contact_edit_view(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id)
    if request.method == "POST":
        form = ContactForm(request.POST, request.FILES, instance=contact)
        if form.is_valid():
            contact = form.save(commit=False)
            contact.user = request.user
            contact.save()
            return HttpResponseRedirect(reverse('contacts:detail', args=(contact.id,)))
    else:
        form = ContactForm(instance=contact)
    return render(request, 'pages/contact_create.html', {'form': form})

@login_required(login_url="/login/")
def postcard_view(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id)
    send_postcard(contact)
    return HttpResponseRedirect(reverse('contacts:detail', args=(contact.id,)))
